#!/usr/bin/env python3
"""
new_contacts_daily_log.py — AXISKEY account
Append-only update of the hand-curated data/exports/axiskey_new_contacts_log.xlsx.
Fetches contacts created (dateAdded) in the target range, bucketed by the
location's own local timezone (not UTC), and:
  - Appends any not already present (by date+name) to the "New Contacts"
    sheet, most recent on top. Never touches existing rows — manual
    deletions/fixes you've made stay gone/fixed.
  - Updates the Leads column for the affected day(s) on the "Joncarlo",
    "Alex", "Stormer" tabs (if present), leaving every other column
    (Discovery Calls, Due Diligence Calls, Proposal Sent, Agreement
    Signed) and row exactly as you left it.
  - Rechecks every row still blank on Source or Owner against GHL (fields
    sometimes get filled in after contact creation) and fills them in if
    now available — never overwrites a value already in the sheet.
  - Fully rebuilds "Contacts per owner" every run (it's derived data with
    no manual columns, so full regen is safe) — a week-by-week count per
    owner, Week 1 = Aug 7-13 2026, plus a stacked column chart (X axis =
    week, Y axis = new contacts).
data/new_contacts_log.json is a separate running record of every contact
ever fetched (for exclusion-purge and dedup bookkeeping) — it is not the
Excel file's source of truth anymore; the saved .xlsx is.

Read-only against GHL: search query only, no CRM data is changed.

Usage:
  python3 new_contacts_daily_log.py                 # yesterday only (the daily trigger)
  python3 new_contacts_daily_log.py --date 2026-08-09        # one specific day
  python3 new_contacts_daily_log.py --start 2026-08-07 --end 2026-08-10  # backfill a range
"""

import argparse
import http.client
import json
import ssl
import time
import urllib.parse
from collections import Counter
from datetime import datetime, timedelta, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Font


# ─── 1. ARGS ────────────────────────────────────────────────────────────────
# No arguments = yesterday only (local time) — this is the daily "update" trigger.

parser = argparse.ArgumentParser()
parser.add_argument("--date", help="Single day to fetch, YYYY-MM-DD (local)")
parser.add_argument("--start", help="Backfill range start, YYYY-MM-DD (local, inclusive)")
parser.add_argument("--end", help="Backfill range end, YYYY-MM-DD (local, inclusive). Defaults to --start.")
args = parser.parse_args()


# ─── 2. LOAD CREDENTIALS ──────────────────────────────────────────────────

def load_env(path):
    result = {}
    for line in Path(path).read_text().splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            key, _, val = line.partition("=")
            result[key.strip()] = val.strip()
    return result

env         = load_env(Path(__file__).parent / ".env")
TOKEN       = env["GHL_TOKEN_AXISKEY"]
LOCATION_ID = env["GHL_LOCATION_ID_AXISKEY"]

HEADERS = {
    "Authorization": f"Bearer {TOKEN}",
    "Version":       "2021-07-28",
    "Accept":        "application/json",
    "Content-Type":  "application/json",
}


# ─── 3. API HELPERS ───────────────────────────────────────────────────────
# GHL occasionally returns a transient 401 with message "Command timed out"
# (seen on /users/ specifically) that has nothing to do with the token being
# bad — retrying once or twice clears it.

def _request(method, url_path, body=None, retries=2):
    for attempt in range(retries + 1):
        conn = http.client.HTTPSConnection(
            "services.leadconnectorhq.com",
            context=ssl.create_default_context(),
        )
        conn.request(method, url_path, body=body, headers=HEADERS)
        resp = conn.getresponse()
        raw = resp.read()
        if resp.status == 200:
            return json.loads(raw)
        if attempt < retries and b"timed out" in raw.lower():
            time.sleep(1)
            continue
        raise Exception(f"HTTP {resp.status} on {url_path}: {raw.decode()}")

def ghl_get(path, params=None):
    url_path = path
    if params:
        url_path += "?" + urllib.parse.urlencode(params)
    return _request("GET", url_path)

def ghl_post(path, body):
    return _request("POST", path, body=json.dumps(body))


# ─── 4. LOCATION TIMEZONE ──────────────────────────────────────────────────
# GHL stores dateAdded in UTC, but "which day" a contact belongs to is a
# local-calendar-day question. Confirmed 2026-08-11: America/New_York for
# AXISKEY. Fetched fresh each run rather than hardcoded.

print("Fetching location timezone...")
location = ghl_get(f"/locations/{LOCATION_ID}").get("location", {})
LOCAL_TZ = ZoneInfo(location.get("timezone") or "UTC")
print(f"  Timezone: {LOCAL_TZ}")
print()


# ─── 5. RESOLVE THE DATE RANGE FOR THIS RUN ────────────────────────────────

if args.start:
    START_DATE = args.start
    END_DATE   = args.end or args.start
elif args.date:
    START_DATE = END_DATE = args.date
else:
    yesterday_local = datetime.now(LOCAL_TZ) - timedelta(days=1)
    START_DATE = END_DATE = yesterday_local.strftime("%Y-%m-%d")

local_start = datetime.strptime(START_DATE, "%Y-%m-%d").replace(tzinfo=LOCAL_TZ)
local_end   = datetime.strptime(END_DATE, "%Y-%m-%d").replace(tzinfo=LOCAL_TZ) + timedelta(days=1) - timedelta(milliseconds=1)
range_start = local_start.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3] + "Z"
range_end   = local_end.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3] + "Z"


# ─── 6. FIND THE "Source" CUSTOM FIELD ID ─────────────────────────────────
# Looked up by name each run (rather than hardcoded) so this keeps working
# if the field is ever recreated. Confirmed 2026-08-11: id SJuKzd9iY0Ib0OUjyD5G,
# fieldKey contact.source_new, model=contact, options MGL/SGL/Referral/
# Partnership/Client.

print("Fetching custom fields...")
custom_fields = ghl_get(f"/locations/{LOCATION_ID}/customFields").get("customFields", [])
source_field = next(
    (f for f in custom_fields if f.get("model") == "contact" and f.get("name", "").lower() == "source"),
    None,
)
if not source_field:
    raise Exception('No contact-level custom field named "Source" found.')
SOURCE_FIELD_ID = source_field["id"]
print(f'  Found "Source" field: {SOURCE_FIELD_ID}')
print()


# ─── 7. FETCH USERS (FOR OWNER NAMES) ─────────────────────────────────────
# Contacts only carry an assignedTo user ID, not a name. Build id → name
# once so every contact's Owner column can be a name.

print("Fetching users...")
users = ghl_get("/users/", {"locationId": LOCATION_ID}).get("users", [])
user_map = {u["id"]: u.get("name") or f'{u.get("firstName","")} {u.get("lastName","")}'.strip() for u in users}
print(f"  Found {len(user_map)} users")
print()


# ─── 8. FETCH CONTACTS CREATED IN THE DATE RANGE ──────────────────────────
# POST /contacts/search supports a dateAdded range filter directly, sorted
# ascending, paginated with the searchAfter cursor GHL returns per record.

print(f"Fetching contacts created {START_DATE} → {END_DATE}...")
all_contacts, search_after = [], None
while True:
    body = {
        "locationId": LOCATION_ID,
        "pageLimit": 100,
        "filters": [
            {"field": "dateAdded", "operator": "range", "value": {"gte": range_start, "lte": range_end}}
        ],
        "sort": [{"field": "dateAdded", "direction": "asc"}],
    }
    if search_after:
        body["searchAfter"] = search_after

    batch = ghl_post("/contacts/search", body).get("contacts", [])
    all_contacts.extend(batch)
    print(f"  Batch: {len(batch)} records")
    if len(batch) < 100:
        break
    search_after = batch[-1]["searchAfter"]

print(f"  Total fetched: {len(all_contacts)}")
print()


# ─── 9. BUILD ROWS FOR THIS RUN ────────────────────────────────────────────
# The UTC search window is padded to cover the local date range, so re-check
# each contact's LOCAL day and drop any that land just outside START/END.
#
# Excluded contacts (test entries / dummy data, per operator 2026-08-12):
# any name containing "test" (case-insensitive), or the contact literally
# named "Joncarlo Tamayo" (not to be confused with the rep of the same
# name, who shows up in the Owner column and is untouched by this).

EXCLUDE_NAME_CONTAINS = ["test"]
EXCLUDE_NAME_EXACT    = ["joncarlo tamayo"]

def is_excluded(name):
    lname = name.lower()
    return any(s in lname for s in EXCLUDE_NAME_CONTAINS) or lname in EXCLUDE_NAME_EXACT

fetched_rows = {}
excluded_count = 0
for c in all_contacts:
    created_local = datetime.fromisoformat(c["dateAdded"].replace("Z", "+00:00")).astimezone(LOCAL_TZ)
    local_date = created_local.strftime("%Y-%m-%d")
    if not (START_DATE <= local_date <= END_DATE):
        continue

    name = f'{c.get("firstName") or ""} {c.get("lastName") or ""}'.strip() or c.get("email") or "(no name)"
    if is_excluded(name):
        excluded_count += 1
        continue

    source_value = ""
    for cf in c.get("customFields", []):
        if cf.get("id") == SOURCE_FIELD_ID:
            source_value = cf.get("value", "")
            break

    owner = user_map.get(c.get("assignedTo"), "") if c.get("assignedTo") else ""

    fetched_rows[c["id"]] = {
        "id": c["id"],
        "created_local": created_local.isoformat(),
        "date": local_date,
        "name": name,
        "source": source_value,
        "owner": owner,
    }


# ─── 10. MERGE INTO THE PERSISTENT LOG ─────────────────────────────────────
# Keyed by contact ID so reruns (same day twice, or overlapping backfills)
# never create duplicate rows — they just overwrite with the latest values.

log_path = Path(__file__).parent / "data" / "new_contacts_log.json"
log_path.parent.mkdir(parents=True, exist_ok=True)

if log_path.exists():
    log_by_id = {r["id"]: r for r in json.loads(log_path.read_text())}
else:
    log_by_id = {}

purged = [r for r in log_by_id.values() if is_excluded(r["name"])]
for r in purged:
    del log_by_id[r["id"]]
if purged:
    excluded_count += len(purged)
    print(f"Purged {len(purged)} previously-logged excluded contact(s): {', '.join(r['name'] for r in purged)}")

new_count = sum(1 for cid in fetched_rows if cid not in log_by_id)
log_by_id.update(fetched_rows)

log_path.write_text(json.dumps(list(log_by_id.values()), indent=2))
print(f"Log updated: {new_count} new contact(s) added for {START_DATE}"
      + (f" → {END_DATE}" if END_DATE != START_DATE else "")
      + f", {excluded_count} excluded (test/Joncarlo Tamayo)."
      + f" Log now covers {len(log_by_id)} contacts total.")
print()


# ─── 11. APPEND-ONLY UPDATE OF THE SAVED WORKBOOK ──────────────────────────
# The workbook is a hand-curated file (rows deleted/fixed, sheets renamed,
# owner tracker tabs with manually-entered Discovery/Diligence/Proposal/
# Agreement data) — added 2026-08-12. This script NEVER rebuilds it. It only:
#   1. Appends brand-new contacts to "New Contacts" (matched by date+name
#      against what's already there, so nothing already deleted comes back).
#   2. Updates the Leads count for the affected day(s) on the "Joncarlo",
#      "Alex", "Stormer" owner tabs, if those tabs exist — inserting a new
#      Day row if needed, or correcting Leads in place if the day is already
#      there. The Discovery/Diligence/Proposal/Agreement columns and every
#      other row are left completely untouched.
# "By Source" is a legacy pivot sheet — no longer auto-updated; edit by
# hand or ask for a one-off rebuild if needed. "Contacts per owner" IS
# rebuilt automatically — see section 12 below.

exports_dir = Path(__file__).parent / "data" / "exports"
exports_dir.mkdir(parents=True, exist_ok=True)
output_path = exports_dir / "axiskey_new_contacts_log.xlsx"

if output_path.exists():
    wb = load_workbook(output_path)
    ws = wb["New Contacts"]
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "New Contacts"
    ws.append(["Date", "Name", "Source", "Owner"])
    for cell in ws[1]:
        cell.font = Font(bold=True)


def insert_sorted_by_date(sheet, date_str, row_values, date_col=1):
    """Insert row_values at the row where date_str actually belongs, keeping
    the sheet sorted newest-first — NOT always at row 2. Bug fixed 2026-08-18:
    always inserting at row 2 assumes every existing row is older than the
    new one, which is only true for a plain same-day run. It breaks the
    moment a backfill adds date(s) older than something already appended
    earlier in the same session (e.g. default "update" adds day N, then a
    backfill for day N-3..N-1 runs after — those got shoved above day N
    instead of below it). This scans for the first existing row whose date
    is <= date_str and inserts there (falling to the very bottom if
    date_str is older than everything present)."""
    insert_at = sheet.max_row + 1
    for i in range(2, sheet.max_row + 1):
        existing_date = sheet.cell(row=i, column=date_col).value
        if existing_date is not None and existing_date <= date_str:
            insert_at = i
            break
    sheet.insert_rows(insert_at)
    for col, val in enumerate(row_values, start=1):
        sheet.cell(row=insert_at, column=col, value=val)


existing_keys = {(row[0].value, row[1].value) for row in ws.iter_rows(min_row=2)}

new_rows = [r for r in fetched_rows.values() if (r["date"], r["name"]) not in existing_keys]
new_rows.sort(key=lambda r: r["created_local"])  # oldest first, so same-day new rows land newest-on-top of each other

for r in new_rows:
    insert_sorted_by_date(ws, r["date"], [r["date"], r["name"], r["source"], r["owner"]])

print(f"Appended {len(new_rows)} new row(s) to New Contacts (skipped {len(fetched_rows) - len(new_rows)} already present).")

OWNER_TABS = {"Joncarlo Tamayo": "Joncarlo", "Alex Zinny": "Alex", "Stormer Santana": "Stormer", "Cole Lytle": "Cole"}
all_rows_now = [row for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]
leads_by_owner_day = Counter((row[3], row[0]) for row in all_rows_now)
affected_dates = sorted({r["date"] for r in new_rows})

for owner, tab in OWNER_TABS.items():
    if tab not in wb.sheetnames:
        continue
    ots = wb[tab]
    for d in affected_dates:
        count = leads_by_owner_day.get((owner, d), 0)
        day_rows = {ots.cell(row=i, column=1).value: i for i in range(2, ots.max_row + 1)}
        if d in day_rows:
            ots.cell(row=day_rows[d], column=2, value=count)
        else:
            insert_sorted_by_date(ots, d, [d, count])
    print(f"  {tab}: Leads updated for {', '.join(affected_dates) if affected_dates else '(no new days)'}")


# ─── 12. BACKFILL BLANK Source/Owner ON PAST CONTACTS ──────────────────────
# GHL contacts sometimes get Source or Owner set *after* creation (a rep
# fills the field in, or gets assigned, a bit later). Every "update" run,
# recheck any row in New Contacts still blank on either field and see if
# GHL now has a value. Only ever fills a blank — never overwrites a value
# already in the sheet, manual or automated. Added 2026-08-18.

id_by_date_name = {(r["date"], r["name"]): r["id"] for r in log_by_id.values()}

blank_rows = [
    (i, row[0], row[1])
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2)
    if row[0] and (not row[2] or not row[3])
]

filled, not_found = 0, 0
for row_idx, row_date, row_name in blank_rows:
    cid = id_by_date_name.get((row_date, row_name))
    if not cid:
        continue
    try:
        contact = ghl_get(f"/contacts/{cid}").get("contact", {})
    except Exception:
        not_found += 1  # contact deleted/merged in GHL since it was logged — skip, don't crash the run
        continue
    time.sleep(0.1)

    if not ws.cell(row=row_idx, column=3).value:
        for cf in contact.get("customFields", []):
            if cf.get("id") == SOURCE_FIELD_ID and cf.get("value"):
                ws.cell(row=row_idx, column=3, value=cf["value"])
                filled += 1
                break

    if not ws.cell(row=row_idx, column=4).value:
        owner_name = user_map.get(contact.get("assignedTo"))
        if owner_name:
            ws.cell(row=row_idx, column=4, value=owner_name)
            filled += 1

print(f"Checked {len(blank_rows)} row(s) with blank Source/Owner; filled {filled} field(s) newly available in GHL"
      + (f" ({not_found} no longer found in GHL, skipped)" if not_found else "") + ".")


# ─── 13. REBUILD "Contacts per owner" (fully derived — safe to regenerate) ─
# Unlike New Contacts / the owner tabs, this sheet has no manually-entered
# columns, so it's rebuilt from scratch from the current New Contacts data
# every run rather than patched — simpler and always correct. Bucketed by
# week (Week 1 = Aug 7–13, the first week of tracking), not by day.
#
# Three tables + charts, added 2026-08-14 (Owner and Owner×Source per
# operator request that day; Source was already weekly from the previous
# rebuild — the old day-bucketed standalone "By Source" sheet is retired,
# folded in here beside the Owner chart instead):
#   1. Owner per week       (top-left)
#   2. Source per week      (top-right, beside #1's chart)
#   3. Owner × Source per week (below both — one row per week+owner, source
#      columns; chart categories are a helper "Week · Owner" label column
#      to the right of the table, since openpyxl's multi-level category
#      axis support is unreliable across Excel versions)

WEEK1_START = datetime.strptime("2026-08-07", "%Y-%m-%d")

def week_label(date_str):
    d = datetime.strptime(date_str, "%Y-%m-%d")
    week_idx = (d - WEEK1_START).days // 7
    start = WEEK1_START + timedelta(days=week_idx * 7)
    end = start + timedelta(days=6)
    return week_idx, f"{start.strftime('%b %-d')} - {end.strftime('%b %-d')}"

all_rows_now = [row for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]  # re-read post-append
owners_present  = sorted({row[3] or "(unassigned)" for row in all_rows_now})
sources_present = sorted({row[2] or "(none)"       for row in all_rows_now})

week_owner_counts  = Counter()  # (week_idx, owner)  -> count
week_source_counts = Counter()  # (week_idx, source) -> count
week_owner_source_counts = Counter()  # (week_idx, owner, source) -> count
week_labels = {}

for date_, name, source, owner in all_rows_now:
    idx, label = week_label(date_)
    week_labels[idx] = label
    o, s = owner or "(unassigned)", source or "(none)"
    week_owner_counts[(idx, o)] += 1
    week_source_counts[(idx, s)] += 1
    week_owner_source_counts[(idx, o, s)] += 1

weeks_sorted = sorted(week_labels)
n_weeks = len(weeks_sorted)

if "Contacts per owner" in wb.sheetnames:
    del wb["Contacts per owner"]
if "By Source" in wb.sheetnames:
    del wb["By Source"]
cpo = wb.create_sheet("Contacts per owner")


def write_week_table(start_row, start_col, categories, counts_dict):
    """Week (rows) × categories (cols) count table with a TOTAL row/col.
    Returns (header_row, last_week_row, total_row, total_col)."""
    ws_ = cpo
    ws_.cell(row=start_row, column=start_col, value="Week")
    for i, cat in enumerate(categories):
        ws_.cell(row=start_row, column=start_col + 1 + i, value=cat)
    total_col = start_col + 1 + len(categories)
    ws_.cell(row=start_row, column=total_col, value="Total")
    for c in range(start_col, total_col + 1):
        ws_.cell(row=start_row, column=c).font = Font(bold=True)

    for r, idx in enumerate(weeks_sorted):
        row = start_row + 1 + r
        ws_.cell(row=row, column=start_col, value=week_labels[idx])
        row_counts = [counts_dict.get((idx, cat), 0) for cat in categories]
        for i, val in enumerate(row_counts):
            ws_.cell(row=row, column=start_col + 1 + i, value=val)
        ws_.cell(row=row, column=total_col, value=sum(row_counts))

    total_row = start_row + 1 + n_weeks
    ws_.cell(row=total_row, column=start_col, value="TOTAL")
    totals = [sum(counts_dict.get((idx, cat), 0) for idx in weeks_sorted) for cat in categories]
    for i, val in enumerate(totals):
        ws_.cell(row=total_row, column=start_col + 1 + i, value=val)
    ws_.cell(row=total_row, column=total_col, value=sum(totals))
    for c in range(start_col, total_col + 1):
        ws_.cell(row=total_row, column=c).font = Font(bold=True)

    for c in range(start_col, total_col + 1):
        ws_.column_dimensions[ws_.cell(row=start_row, column=c).column_letter].width = 16

    return start_row, start_row + n_weeks, total_row, total_col


def add_week_chart(title, header_row, last_week_row, data_min_col, data_max_col, cat_col, anchor_cell):
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = title
    chart.y_axis.title = "New contacts"
    chart.x_axis.title = "Week"
    data = Reference(cpo, min_col=data_min_col, max_col=data_max_col, min_row=header_row, max_row=last_week_row)
    cats = Reference(cpo, min_col=cat_col, max_col=cat_col, min_row=header_row + 1, max_row=last_week_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height, chart.width = 10, 16
    cpo.add_chart(chart, anchor_cell)
    return chart


# --- Table 1: Owner per week (top-left) ---
o_header, o_last_week, o_total_row, o_total_col = write_week_table(1, 1, owners_present, week_owner_counts)
charts_anchor_row = o_total_row + 3
add_week_chart("New Contacts per Owner per Week", o_header, o_last_week, 2, o_total_col - 1, 1, f"A{charts_anchor_row}")

# --- Table 2: Source per week (top-right, beside Table 1's chart) ---
src_start_col = o_total_col + 2
s_header, s_last_week, s_total_row, s_total_col = write_week_table(1, src_start_col, sources_present, week_source_counts)
add_week_chart(
    "New Contacts per Source per Week", s_header, s_last_week,
    src_start_col + 1, s_total_col - 1, src_start_col, f"{cpo.cell(row=1, column=src_start_col).column_letter}{charts_anchor_row}",
)

# --- Table 3: Owner × Source per week (below both, full width) ---
# Rows grouped by owner (real reps first, "(unassigned)" last if present),
# then week within each owner's block — that makes each owner's rows
# contiguous, so its chart can reference a plain rectangular range instead
# of needing a helper "Week · Owner" label column. One chart per real
# owner (not "(unassigned)" — nothing to chart there per operator request
# 2026-08-14), stacked vertically below the table.
os_start_row = charts_anchor_row + 22  # clear both charts above (height=10cm ≈ ~19 rows)
cpo.cell(row=os_start_row - 1, column=1, value="Owner × Source per week").font = Font(bold=True, size=12)

headers = ["Week", "Owner"] + sources_present + ["Total"]
for i, h in enumerate(headers):
    cpo.cell(row=os_start_row, column=1 + i, value=h).font = Font(bold=True)

real_owners = [o for o in owners_present if o != "(unassigned)"]
owners_order = real_owners + (["(unassigned)"] if "(unassigned)" in owners_present else [])

owner_blocks = {}  # owner -> (first_row, last_row)
row = os_start_row + 1
for owner in owners_order:
    block_first_row = row
    for idx in weeks_sorted:
        row_counts = [week_owner_source_counts.get((idx, owner, s), 0) for s in sources_present]
        cpo.cell(row=row, column=1, value=week_labels[idx])
        cpo.cell(row=row, column=2, value=owner)
        for i, val in enumerate(row_counts):
            cpo.cell(row=row, column=3 + i, value=val)
        cpo.cell(row=row, column=2 + len(sources_present) + 1, value=sum(row_counts))
        row += 1
    owner_blocks[owner] = (block_first_row, row - 1)
os_last_row = row - 1

for c in range(1, len(headers) + 1):
    cpo.column_dimensions[cpo.cell(row=os_start_row, column=c).column_letter].width = 16

chart_row = os_last_row + 3
for owner in real_owners:
    first_row, last_row = owner_blocks[owner]
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = f"{owner} — Contacts per Source per Week"
    chart.y_axis.title = "New contacts"
    chart.x_axis.title = "Week"
    # No header row is adjacent to this owner's block (other owners' rows
    # sit between it and the shared header), so titles_from_data can't be
    # used here — series titles are set explicitly from sources_present
    # instead, sourcing only this owner's contiguous row range.
    data = Reference(cpo, min_col=3, max_col=2 + len(sources_present), min_row=first_row, max_row=last_row)
    cats = Reference(cpo, min_col=1, max_col=1, min_row=first_row, max_row=last_row)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    for i, s_name in enumerate(sources_present):
        chart.series[i].tx = SeriesLabel(v=s_name)
    chart.height, chart.width = 8, 14
    cpo.add_chart(chart, f"A{chart_row}")
    chart_row += 18  # clear each chart (height=8cm ≈ ~15 rows) before the next

print(f"Rebuilt Contacts per owner: {n_weeks} week(s), owners: {', '.join(owners_present)}, sources: {', '.join(sources_present)}")

wb.save(output_path)
print(f"Saved {output_path}")
