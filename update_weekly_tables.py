#!/usr/bin/env python3
"""
update_weekly_tables.py — AXISKEY account

Rebuilds the "Weekly by Rep" sheet in axiskey_new_contacts_log.xlsx: for each
tracked rep, a week-by-stage count matrix and the matching stage-to-stage
conversion matrix.

Source is the six "Date Entered - <Stage>" opportunity DATE custom fields,
the same source of truth as update_funnel.py and build_dashboard.py's
Weekly Rocks.

READ THE CAVEAT: the conversion tables divide stage entries that happened in
the same week, but a lead and its discovery call usually fall in different
weeks, so a stage can exceed 100% of the stage "above" it. These are weekly
throughput ratios, not cohort conversion. Only the Total column is a true
conversion rate. The caveat is written into the sheet too, because the sheet
outlives the conversation that produced it.

Only ever touches the "Weekly by Rep" sheet — every other sheet (including
the hand-curated New Contacts and the per-rep activity tabs) is left alone.

Read-only against GHL: GET requests only.

Run with:  python3 update_weekly_tables.py
"""

import http.client
import json
import ssl
import time
import urllib.parse
from collections import Counter
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ─── CONFIG ────────────────────────────────────────────────────────────────

WEEK0_START = date(2026, 7, 29)   # partial week: Jul 29–31
WEEK1_START = date(2026, 8, 1)    # first full 7-day week

REPS = ["Alex Zinny", "Joncarlo Tamayo"]
# Reps shown as a totals-only block (no week-by-week breakdown), per operator
# request 2026-08-25.
TOTALS_ONLY_REPS = ["Stormer Santana"]
SHORT = {"Alex Zinny": "Alex", "Joncarlo Tamayo": "Joncarlo", "Stormer Santana": "Stormer"}
TRACKED = REPS + TOTALS_ONLY_REPS

STAGES = ["New Lead", "Discovery Call", "Strategy Call", "Proposal Sent", "Agreement Signed"]

FIELD_DATE_STAGES = {
    "YOUfzDu5jq9T3EpsdtgL": "New Lead",
    "fb5FWif6GUyl4c3E60bR": "Discovery Call",
    "aVXVp6kynBp7taut7l3Z": "Strategy Call",
    "hiccRLHd3sqdrYPErOkc": "Proposal Sent",
    "Qru4091H66VTPvH2rQKO": "Agreement Signed",
}

CHARCOAL = "36454F"
BAND     = "F2F2F2"

# ─── CREDENTIALS & API ─────────────────────────────────────────────────────

def load_env(path):
    out = {}
    for line in Path(path).read_text().splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            k, _, v = line.partition("=")
            out[k.strip()] = v.strip()
    return out

env = load_env(Path(__file__).parent / ".env")
TOKEN, LOCATION_ID = env["GHL_TOKEN_AXISKEY"], env["GHL_LOCATION_ID_AXISKEY"]
HEADERS = {"Authorization": f"Bearer {TOKEN}", "Version": "2021-07-28", "Accept": "application/json"}


def ghl_get(path, params=None, retries=2):
    url = path + ("?" + urllib.parse.urlencode(params) if params else "")
    for attempt in range(retries + 1):
        conn = http.client.HTTPSConnection(
            "services.leadconnectorhq.com", context=ssl.create_default_context(), timeout=30)
        conn.request("GET", url, headers=HEADERS)
        resp = conn.getresponse()
        raw = resp.read()
        if resp.status == 200:
            return json.loads(raw)
        if attempt < retries and b"timed out" in raw.lower():
            time.sleep(1)
            continue
        raise Exception(f"HTTP {resp.status} on {url}: {raw.decode()[:300]}")


# ─── WEEK BUCKETS (auto-extend to today) ───────────────────────────────────

def build_weeks(today):
    weeks = [("W0", WEEK0_START, WEEK1_START - timedelta(days=1))]
    i, start = 1, WEEK1_START
    while start <= today:
        end = start + timedelta(days=6)
        weeks.append((f"W{i}", start, min(end, today) if end > today else end))
        start, i = end + timedelta(days=1), i + 1
    return weeks


def week_label(tag, s, e):
    partial = (e - s).days < 6
    span = f"{s.strftime('%b %-d')}–{e.strftime('%-d')}"
    return f"{tag} ({span})" + (" *" if partial else "")


# ─── FETCH ─────────────────────────────────────────────────────────────────

print("Fetching users...")
users = ghl_get("/users/", {"locationId": LOCATION_ID}).get("users", [])
user_map = {u["id"]: (u.get("name") or f'{u.get("firstName","")} {u.get("lastName","")}'.strip()) for u in users}

print("Fetching opportunities...")
opps, page = [], 1
while True:
    batch = ghl_get("/opportunities/search",
                    {"location_id": LOCATION_ID, "limit": 100, "page": page}).get("opportunities", [])
    opps.extend(batch)
    if len(batch) < 100:
        break
    page += 1
print(f"  {len(opps)} opportunities")

today = datetime.now(timezone.utc).date()
WEEKS = build_weeks(today)

counts = Counter()   # (rep, week_tag, stage) -> n
future = 0
for opp in opps:
    owner = user_map.get(opp.get("assignedTo"))
    if owner not in TRACKED:
        continue
    for cf in (opp.get("customFields") or []):
        stage = FIELD_DATE_STAGES.get(cf.get("id"))
        ms = cf.get("fieldValueDate")
        if not stage or not ms:
            continue
        d = datetime.fromtimestamp(ms / 1000, tz=timezone.utc).date()
        if d > today:
            future += 1
            continue
        for tag, s, e in WEEKS:
            if s <= d <= e:
                counts[(owner, tag, stage)] += 1
                break

if future:
    print(f"  ⚠ {future} future-dated stage entries skipped (data-entry errors — fix in GHL)")


# ─── WRITE THE SHEET ───────────────────────────────────────────────────────

path = Path(__file__).parent / "data" / "exports" / "axiskey_new_contacts_log.xlsx"
wb = load_workbook(path)
if "Weekly by Rep" in wb.sheetnames:
    del wb["Weekly by Rep"]
ws = wb.create_sheet("Weekly by Rep")

labels = [week_label(t, s, e) for t, s, e in WEEKS]
ncols = len(labels)

ws["A1"] = f"Weekly stage activity by rep — {WEEK0_START.strftime('%b %-d')} to {today.strftime('%b %-d, %Y')}"
ws["A1"].font = Font(bold=True, size=14, color=CHARCOAL)

row = 3


def write_table(title, rep, as_pct):
    global row
    ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=12, color=CHARCOAL)
    row += 1

    ws.cell(row=row, column=1, value="Stage").font = Font(bold=True)
    for i, lab in enumerate(labels):
        c = ws.cell(row=row, column=2 + i, value=lab)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    c = ws.cell(row=row, column=2 + ncols, value="Total")
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal="center")
    for col in range(1, ncols + 3):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=BAND)
    row += 1

    for si, stage in enumerate(STAGES):
        ws.cell(row=row, column=1, value=stage)
        for i, (tag, _, _) in enumerate(WEEKS):
            cell = ws.cell(row=row, column=2 + i)
            if not as_pct:
                cell.value = counts.get((rep, tag, stage), 0)
                cell.alignment = Alignment(horizontal="center")
            elif si > 0:
                prev = counts.get((rep, tag, STAGES[si - 1]), 0)
                cur = counts.get((rep, tag, stage), 0)
                if prev:
                    cell.value = cur / prev
                    cell.number_format = "0%"
                cell.alignment = Alignment(horizontal="center")

        tcell = ws.cell(row=row, column=2 + ncols)
        if not as_pct:
            tcell.value = sum(counts.get((rep, t, stage), 0) for t, _, _ in WEEKS)
        elif si > 0:
            pt = sum(counts.get((rep, t, STAGES[si - 1]), 0) for t, _, _ in WEEKS)
            ct = sum(counts.get((rep, t, stage), 0) for t, _, _ in WEEKS)
            if pt:
                tcell.value = ct / pt
                tcell.number_format = "0.0%"
        tcell.font = Font(bold=True)
        tcell.alignment = Alignment(horizontal="center")
        row += 1

    row += 2


for rep in REPS:
    write_table(f"{SHORT[rep]} — counts", rep, as_pct=False)
    write_table(f"{SHORT[rep]} — conversion from previous stage", rep, as_pct=True)


def write_totals_table(rep):
    """Totals-only block: no week columns, just count and conversion per stage."""
    global row
    ws.cell(row=row, column=1, value=f"{SHORT[rep]} — totals ({WEEK0_START.strftime('%b %-d')} to {today.strftime('%b %-d')})"
            ).font = Font(bold=True, size=12, color=CHARCOAL)
    row += 1

    for i, h in enumerate(["Stage", "Count", "Conversion from previous"]):
        c = ws.cell(row=row, column=1 + i, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor=BAND)
        if i:
            c.alignment = Alignment(horizontal="center")
    row += 1

    totals = [sum(counts.get((rep, t, st), 0) for t, _, _ in WEEKS) for st in STAGES]
    for i, stage in enumerate(STAGES):
        ws.cell(row=row, column=1, value=stage)
        c = ws.cell(row=row, column=2, value=totals[i])
        c.alignment = Alignment(horizontal="center")
        pc = ws.cell(row=row, column=3)
        if i > 0 and totals[i - 1]:
            pc.value = totals[i] / totals[i - 1]
            pc.number_format = "0.0%"
        pc.alignment = Alignment(horizontal="center")
        row += 1
    row += 2


for rep in TOTALS_ONLY_REPS:
    write_totals_table(rep)

ws.cell(row=row, column=1, value="* Partial week (fewer than 7 days) — counts are not comparable to full weeks.")
ws.cell(row=row, column=1).font = Font(italic=True, size=9)
row += 1
ws.cell(row=row, column=1, value=(
    "CAVEAT: weekly conversion divides stage entries occurring in the same week, but a lead and its "
    "discovery call usually fall in different weeks — so a stage can exceed 100% of the stage above it. "
    "These are weekly throughput ratios, not cohort conversion. Only the Total column is a true "
    "conversion rate. Blank means the previous stage had no entries that week."))
ws.cell(row=row, column=1).font = Font(italic=True, size=9, color="B00000")
row += 1
ws.cell(row=row, column=1, value=f"Generated {today} from GHL 'Date Entered' stage fields. Rebuilt in full on each run.")
ws.cell(row=row, column=1).font = Font(italic=True, size=9)

ws.column_dimensions["A"].width = 26
for i in range(ncols + 1):
    ws.column_dimensions[get_column_letter(2 + i)].width = 15
ws.freeze_panes = "B1"

wb.save(path)
print(f'\nWrote "Weekly by Rep" ({len(WEEKS)} weeks x {len(STAGES)} stages, {len(REPS)} reps) to {path}')
