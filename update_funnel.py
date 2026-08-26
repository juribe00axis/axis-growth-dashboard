#!/usr/bin/env python3
"""
update_funnel.py — AXISKEY account
Rebuilds the "Funnel" sheet in axiskey_new_contacts_log.xlsx from GHL's
"Date Entered - <Stage>" opportunity custom fields (same source of truth
as build_dashboard.py's Weekly Rocks) — counts every stage entry since
SINCE, per rep. Automated 2026-08-25; before that the counts were typed
in by hand under the mistaken belief they weren't pullable.

Only ever touches the "Funnel" sheet — every other sheet in the workbook
is left alone (same append-only spirit as new_contacts_daily_log.py).

Read-only against GHL: GET requests only, no CRM data is changed.

Run with:  python3 update_funnel.py
"""

import http.client
import json
import ssl
import time
import urllib.parse
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Font

# ─── CONFIG ────────────────────────────────────────────────────────────────

SINCE = "2026-07-29"

# Reps shown in the funnel, in display order (operator keeps this to Alex +
# Joncarlo for now, 2026-08-25 — Stormer/Cole deliberately not included).
FUNNEL_REPS = ["Alex Zinny", "Joncarlo Tamayo"]
DISPLAY_NAMES = {"Alex Zinny": "Alex", "Joncarlo Tamayo": "Joncarlo"}

# Funnel stages in order. "Agreement Sent" exists in GHL but is not shown
# (matches the dashboard's Weekly Rocks, which also hides it).
FUNNEL_STAGES = ["New Lead", "Discovery Calls", "Strategy Calls", "Proposal Sent", "Agreement Signed"]

# GHL "Date Entered - <Stage>" DATE custom fields on opportunities — the
# true stage-entry dates. Same mapping as build_dashboard.py
# FIELD_DATE_STAGES, validated live 2026-07-27.
FIELD_DATE_STAGES = {
    "YOUfzDu5jq9T3EpsdtgL": "New Lead",
    "fb5FWif6GUyl4c3E60bR": "Discovery Calls",
    "aVXVp6kynBp7taut7l3Z": "Strategy Calls",
    "hiccRLHd3sqdrYPErOkc": "Proposal Sent",
    "Qru4091H66VTPvH2rQKO": "Agreement Signed",
}

BLOCK_HEIGHT = 26  # rows reserved per rep block, table + chart
HELPER_COL   = 8   # column H — chart source data, kept separate from the readable table


# ─── CREDENTIALS & API ─────────────────────────────────────────────────────

def load_env(path):
    result = {}
    for line in Path(path).read_text().splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            key, _, val = line.partition("=")
            result[key.strip()] = val.strip()
    return result

env         = load_env(Path(__file__).parent / ".env")
TOKEN       = env["GHL_TOKEN_AXISKEY"]
LOCATION_ID = env["GHL_LOCATION_ID_AXISKEY"]

HEADERS = {
    "Authorization": f"Bearer {TOKEN}",
    "Version":       "2021-07-28",
    "Accept":        "application/json",
}

def ghl_get(path, params=None, retries=2):
    url = path + ("?" + urllib.parse.urlencode(params) if params else "")
    for attempt in range(retries + 1):
        conn = http.client.HTTPSConnection(
            "services.leadconnectorhq.com",
            context=ssl.create_default_context(), timeout=30,
        )
        conn.request("GET", url, headers=HEADERS)
        resp = conn.getresponse()
        raw = resp.read()
        if resp.status == 200:
            return json.loads(raw)
        if attempt < retries and b"timed out" in raw.lower():
            time.sleep(1)
            continue
        raise Exception(f"HTTP {resp.status} on {url}: {raw.decode()[:300]}")


# ─── FETCH & COUNT STAGE ENTRIES ───────────────────────────────────────────

print("Fetching users...")
users = ghl_get("/users/", {"locationId": LOCATION_ID}).get("users", [])
user_map = {u["id"]: (u.get("name") or f'{u.get("firstName","")} {u.get("lastName","")}'.strip()) for u in users}

print("Fetching all opportunities...")
all_opps, page = [], 1
while True:
    batch = ghl_get("/opportunities/search", {"location_id": LOCATION_ID, "limit": 100, "page": page}).get("opportunities", [])
    all_opps.extend(batch)
    if len(batch) < 100:
        break
    page += 1
print(f"  {len(all_opps)} opportunities")

since_date = datetime.strptime(SINCE, "%Y-%m-%d").date()
today      = datetime.now(timezone.utc).date()

counts = Counter()  # (owner_name, stage) -> count
future_dated_skipped = 0
for opp in all_opps:
    owner = user_map.get(opp.get("assignedTo"))
    if owner not in FUNNEL_REPS:
        continue
    for cf in (opp.get("customFields") or []):
        stage = FIELD_DATE_STAGES.get(cf.get("id"))
        ms = cf.get("fieldValueDate")
        if not stage or not ms:
            continue
        d = datetime.fromtimestamp(ms / 1000, tz=timezone.utc).date()
        # A Date-Entered value after today is a data-entry mistake (wrong
        # year picked), not a real future stage entry — same rule as
        # build_dashboard.py, confirmed 2026-08-18.
        if d > today:
            future_dated_skipped += 1
            continue
        if d >= since_date:
            counts[(owner, stage)] += 1

FUNNEL_DATA = {
    DISPLAY_NAMES[rep]: {stage: counts.get((rep, stage), 0) for stage in FUNNEL_STAGES}
    for rep in FUNNEL_REPS
}

print(f"\nStage entries since {SINCE}:")
for rep, stages in FUNNEL_DATA.items():
    print(f"  {rep}: " + ", ".join(f"{s}={c}" for s, c in stages.items()))
if future_dated_skipped:
    print(f"  ⚠ {future_dated_skipped} future-dated entries skipped (data-entry errors — check in GHL)")


# ─── BUILD THE SHEET ───────────────────────────────────────────────────────

output_path = Path(__file__).parent / "data" / "exports" / "axiskey_new_contacts_log.xlsx"
wb = load_workbook(output_path)

if "Funnel" in wb.sheetnames:
    del wb["Funnel"]
ws = wb.create_sheet("Funnel")

ws["A1"] = f"Sales Funnel — since {SINCE} (updated {today.strftime('%Y-%m-%d')})"
ws["A1"].font = Font(bold=True, size=14)

row_cursor = 3
for rep, stages in FUNNEL_DATA.items():
    r0 = row_cursor
    ws.cell(row=r0, column=1, value=rep).font = Font(bold=True, size=12)

    headers = ["Stage", "Count", "Conv. from Previous", "Conv. from New Lead"]
    for i, h in enumerate(headers):
        ws.cell(row=r0 + 1, column=1 + i, value=h).font = Font(bold=True)

    stage_names = list(stages.keys())
    stage_counts = list(stages.values())
    first_count = stage_counts[0]

    for i, (name, count) in enumerate(zip(stage_names, stage_counts)):
        row = r0 + 2 + i
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=count)
        if i > 0 and stage_counts[i - 1] > 0:
            cell = ws.cell(row=row, column=3, value=count / stage_counts[i - 1])
            cell.number_format = "0.0%"
        if first_count > 0:
            cell = ws.cell(row=row, column=4, value=count / first_count)
            cell.number_format = "0.0%"

    # Hidden-in-plain-sight helper columns feeding the funnel chart: a
    # "Blank" padding series (half the gap to the widest stage, invisible)
    # plus the real "Count" series, stacked — the standard Excel technique
    # for faking a symmetric funnel shape out of a horizontal bar chart.
    n = len(stage_names)
    max_count = max(stage_counts) or 1
    ws.cell(row=r0 + 1, column=HELPER_COL,     value="Stage")
    ws.cell(row=r0 + 1, column=HELPER_COL + 1, value="Blank")
    ws.cell(row=r0 + 1, column=HELPER_COL + 2, value="Count")
    for i, (name, count) in enumerate(zip(stage_names, stage_counts)):
        row = r0 + 2 + i
        ws.cell(row=row, column=HELPER_COL,     value=name)
        ws.cell(row=row, column=HELPER_COL + 1, value=(max_count - count) / 2)
        ws.cell(row=row, column=HELPER_COL + 2, value=count)

    data_first_row = r0 + 2
    data_last_row  = r0 + 1 + n

    chart = BarChart()
    chart.type = "bar"          # horizontal bars
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.gapWidth = 20
    chart.title = f"{rep} Funnel (since {SINCE})"
    chart.legend = None
    chart.y_axis.delete = True                       # hide the count axis (data labels show counts)
    chart.x_axis.delete = False                      # keep stage names visible
    chart.x_axis.scaling.orientation = "maxMin"      # New Lead on top, Agreement Signed on bottom

    data = Reference(ws, min_col=HELPER_COL + 1, max_col=HELPER_COL + 2, min_row=r0 + 1, max_row=data_last_row)
    cats = Reference(ws, min_col=HELPER_COL,      max_col=HELPER_COL,     min_row=data_first_row, max_row=data_last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    chart.series[0].graphicalProperties = GraphicalProperties(noFill=True)          # padding: invisible
    chart.series[1].graphicalProperties = GraphicalProperties(solidFill="4472C4")   # counts: visible
    chart.series[1].dLbls = DataLabelList()
    chart.series[1].dLbls.showVal = True

    chart.height, chart.width = 7.5, 13
    ws.add_chart(chart, f"A{r0 + 2 + n + 1}")

    row_cursor += BLOCK_HEIGHT

widths = {"A": 20, "B": 10, "C": 20, "D": 20}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

wb.save(output_path)
print(f"\nFunnel sheet rebuilt in {output_path}")
